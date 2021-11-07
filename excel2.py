import openpyxl
import glob
import shutil
from natsort import natsorted
import re


# %%
sheet1_row = 1
sheet1_column = 1
sheet2_row =1 
sheet2_column=1
column_list = ['A','B','C','D']
files = glob.glob('C:/Users/USER/Desktop/excel/*.xlsx')

new_file = openpyxl.Workbook()
new_file2 = openpyxl.Workbook()

for name in natsorted(files):#ファイルを上から順番に取得
    before = openpyxl.load_workbook(name)
    sheet1 = before['書面']
    



    number = str(sheet1.cell(2,1).value) #ネジの番号取得
    for row in range(2,48):
        syurui = sheet1.cell(row,2).value #ネジの種類取得
        if syurui == 'ｿｹｯﾄｽｸﾘｭｰ':
            s = sheet1.cell(row,4).value#ネジのサイズ取得
            quanity = sheet1.cell(row,5).value#ネジの数量取得
            
            lists =s.split() #空白判定
            result1 = re.sub(r"\D", "", lists[0]) #空白の中の文字列1
            result2= re.sub(r"\D", "", lists[1]) #空白の後の文字列

            syurui = result1 +"x"+result2

            after1 = new_file.active #新しく作ったファイルを利用可能に
            after2 = new_file2.active
            for row_num in range(1,501):
                after2.row_dimensions[row_num].height = 50
            for column_num in range(1,2):
                after2.column_dimensions[column_list[column_num]].width = 13
            

            sheet2_row += 1

            # if sheet2_row % 11 ==0:
            #     after2.cell(sheet2_row,sheet2_column).value = syurui+'\n'+quanity #値設定
            #     after2.cell(sheet2_row,sheet2_column).alignment = openpyxl.styles.Alignment(wrapText=True)
            #     sheet2_row = 1
            #     sheet2_column += 1

            # else:
            #     after.cell(sheet2_row,sheet2_column).value = syurui+'\n'+quanity
            #     after.cell(sheet2_row,sheet2_column).alignment = openpyxl.styles.Alignment(wrapText=True)
            #     sheet2_row += 1
            
        else:
            # after = new_file.active
            # after.cell(sheet2_row,1).value = '空白だよ～'
            # after.cell(sheet2_row,2).value =   '空白だよ～'
            # sheet2_row += 1
            pass
    shutil.move(name,'C:/Users/USER/Desktop/excel/集計前')

new_file.save('C:/Users/USER/Desktop/excel/集計後/集計1.xlsx')
new_file2.save('C:/Users/USER/Desktop/excel/集計後/集計2.xlsx')
    
        
